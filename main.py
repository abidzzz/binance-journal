from binance.client import Client
import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def get_binance_client(api_key, api_secret):
    return Client(api_key, api_secret)

def fetch_orders(client):
    symbols = ['BNBUSDT', 'DOGEUSDT', 'BONKUSDT', 'NOTUSDT'] # Add all the crypto you trade with
    trades = [["Date and Time (Buy)","Date and Time (Sell)", "Symbol", "Qty (Buy)", "Qty (Sell)", "Buy Price", "Sell Price", "Buy Amount", "Sell Amount", "Profit", "Loss"]]
    for sym in symbols:
        orders = client.get_all_orders(symbol=sym)
        for order in orders:
            order_time = datetime.datetime.fromtimestamp(order['time'] / 1000)
            if order['status'] != "CANCELED":
                if order['side'] == 'BUY' :
                    amount = float(order['price']) * float(order['origQty'])
                    trades.append([
                        order_time.strftime('%a, %d %b %Y %I:%M:%S%p'),
                        '',
                        order['symbol'],
                        order['origQty'],
                        "",
                        order['price'],
                        "",
                        amount,
                        "",
                        "", ""
                    ])
                else:
                    profit = ''
                    loss = ''
                    amount = float(order['price']) * float(order['origQty'])
                    buy_amount = float(trades[-1][7])
                    x  =  buy_amount - amount
                    if buy_amount < amount :
                        profit = x * -1
                    else:
                        loss = x
                    trades[-1][1] = order_time.strftime('%a, %d %b %Y %I:%M:%S%p')
                    trades[-1][4] = order['origQty']
                    trades[-1][6] = order['price']

                    trades[-1][8:] = [
                        amount,
                        profit,
                        loss
                    ]
    return trades

file = open("credentials.txt")
key = file.read().split("\n")

client = get_binance_client(key[0], key[1])

orders = fetch_orders(client)

# Create a workbook and select the active worksheet
wb = Workbook()
ws = wb.active



# Write the orders data to the worksheet
for i, row in enumerate(orders, start=1):
    ws.append(row)
    bold = False
    if i == 1:  # Skip header row
        fill = PatternFill()
        bold = True

    if row[10] == '': # No loss = profit:
        # Green fill for profit
        fill = PatternFill(start_color="d9ead3", end_color="d9ead3", fill_type="solid")
    elif row[9] == '': # No profit = loss:
        # Red fill for loss
        fill = PatternFill(start_color="f4cccc", end_color="f4cccc", fill_type="solid")
    for cell in ws[i]:
        cell.font = Font(name='Consolas', size=12, bold = bold)
        cell.fill = fill

# Save the workbook to a file
wb.save("TradingJournal.xlsx")
print("[*] Saved as TradingJournal.xlsx ")
